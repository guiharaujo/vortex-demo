import os
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import pyodbc
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import re

# ============================================================
# ⚙️  CONFIGURAÇÕES
# ============================================================

DB_SERVER   = os.environ.get("ERP_DB_SERVER", "your-erp-server,1433")
DB_NAME     = os.environ.get("ERP_DB_NAME", "your_erp_database")
DB_USER     = os.environ.get("ERP_DB_USER", "your_db_user")
DB_PASSWORD = os.environ.get("ERP_DB_PASSWORD", "your_db_password")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados", "output.xlsx")

GRUPOS_RECURSIVOS = (124, 126)
GRUPOS_IGNORAR    = (127, 132, 100)

# ============================================================
# 🔌  CONEXÃO
# ============================================================

def conectar():
    print("🔌 Conectando no SAP...")
    conn = pyodbc.connect(
        f"DRIVER={{SQL Server}};"
        f"SERVER={DB_SERVER};"
        f"DATABASE={DB_NAME};"
        f"UID={DB_USER};"
        f"PWD={DB_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    print("✅ Conectado!")
    return conn

# ============================================================
# 📦  BUSCAR DADOS
# ============================================================

def buscar_pas_ativos(conn):
    print("📦 Buscando PAs ativos...")
    query = """
        SELECT
            T.ItemCode,
            T.ItemName,
            T.ItmsGrpCod,
            ISNULL(SUM(W.OnHand), 0) as Estoque
        FROM OITM T WITH(NOLOCK)
        LEFT JOIN OITW W WITH(NOLOCK) ON T.ItemCode = W.ItemCode
            AND W.WhsCode IN ('IA06', 'IE06', 'PD06', 'IE05', 'IE07')
        WHERE T.Canceled = 'N'
          AND T.ItmsGrpCod IN (125, 123)
          AND T.InvntItem = 'Y'
          AND T.frozenFor = 'N'
        GROUP BY T.ItemCode, T.ItemName, T.ItmsGrpCod
        ORDER BY T.ItemCode
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} PAs encontrados")
    return df

def buscar_estoque_insumos(conn):
    print("🔩 Buscando estoque de insumos (IA06 + IE06 + PD06 + IE05 + IE07)...")
    query = """
        SELECT
            T.ItemCode,
            T.ItemName,
            T.ItmsGrpCod,
            ISNULL(SUM(W.OnHand), 0) as Estoque
        FROM OITM T WITH(NOLOCK)
        LEFT JOIN OITW W WITH(NOLOCK) ON T.ItemCode = W.ItemCode
            AND W.WhsCode IN ('IA06', 'IE06', 'PD06', 'IE05', 'IE07')
        WHERE T.Canceled = 'N'
          AND T.InvntItem = 'Y'
        GROUP BY T.ItemCode, T.ItemName, T.ItmsGrpCod
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} itens com estoque consolidado")
    return df

def buscar_bom(conn):
    print("🏗️  Buscando estruturas BOM...")
    query = """
        SELECT B.Father, B.Code AS Insumo, B.Quantity
        FROM ITT1 B WITH(NOLOCK)
        INNER JOIN OITM PA WITH(NOLOCK) ON B.Father=PA.ItemCode AND PA.Canceled='N'
        INNER JOIN OITM INS WITH(NOLOCK) ON B.Code=INS.ItemCode AND INS.Canceled='N'
        ORDER BY B.Father, B.Code
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de BOM")
    return df

def buscar_consumo_3m(conn):
    print("📊 Buscando consumo 3 meses...")
    # Considera apenas NFs de saída:
    #   - não canceladas (H.CANCELED = 'N')
    #   - VINCULADAS A UM PEDIDO DE VENDA (L.BaseType = 17 → ORDR)
    #   - NÃO sendo "1ª NF" de Venda Futura (Comments contém "VD FUTURA") — a 2ª NF (de remessa)
    #     é que vai contar como saída real, evitando duplicidade.
    # NFs sem pedido (demonstração, balcão, cancelamento via devolução) NÃO entram — alinhado ao OTD.
    query = """
        DECLARE @inicio DATE = DATEADD(MONTH, -3, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
        DECLARE @fim    DATE = DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1);
        SELECT L.ItemCode, SUM(L.Quantity) as Consumo3m
        FROM OINV H WITH(NOLOCK)
        INNER JOIN INV1 L WITH(NOLOCK) ON H.DocEntry = L.DocEntry
        INNER JOIN ORDR O WITH(NOLOCK) ON L.BaseEntry = O.DocEntry AND L.BaseType = 17
        WHERE H.DocDate >= @inicio AND H.DocDate < @fim
          AND H.CANCELED = 'N'
          AND ISNULL(H.Comments, '') NOT LIKE '%VD FUTURA%'
        GROUP BY L.ItemCode
    """
    df = pd.read_sql(query, conn)
    df["MediaDiaria"] = df["Consumo3m"] / 90
    df["MediaMensal"] = df["Consumo3m"] / 3
    print(f"✅ Consumo calculado para {len(df)} itens")
    return df

def buscar_mps_ativos(conn):
    """MPs ativas: Mat. Prima (119), Componente (120), Embalagem (121), Consumível (122)."""
    print("🧱 Buscando matérias-primas ativas...")
    query = """
        SELECT
            T.ItemCode,
            T.ItemName,
            T.ItmsGrpCod,
            ISNULL(SUM(W.OnHand), 0) as Estoque
        FROM OITM T WITH(NOLOCK)
        LEFT JOIN OITW W WITH(NOLOCK) ON T.ItemCode = W.ItemCode
            AND W.WhsCode IN ('IA06', 'IE06', 'PD06', 'IE05', 'IE07')
        WHERE T.Canceled = 'N'
          AND T.ItmsGrpCod IN (119, 120, 121, 122)
          AND T.InvntItem = 'Y'
          AND T.frozenFor = 'N'
        GROUP BY T.ItemCode, T.ItemName, T.ItmsGrpCod
        ORDER BY T.ItemCode
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} MPs encontradas")
    return df

def buscar_consumo_3m_mp(conn):
    """Consumo de MP em produção — Goods Issue (OIGE/IGE1) vinculado a OP (BaseType=202).
    3 meses fechados (exclui mês atual). Espelho do buscar_consumo_3m mas pra MP."""
    print("📊 Buscando consumo MP 3 meses (Issue for Production)...")
    query = """
        DECLARE @inicio DATE = DATEADD(MONTH, -3, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
        DECLARE @fim    DATE = DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1);
        SELECT L.ItemCode, SUM(L.Quantity) as Consumo3m
        FROM OIGE H WITH(NOLOCK)
        INNER JOIN IGE1 L WITH(NOLOCK) ON H.DocEntry = L.DocEntry
        WHERE H.DocDate >= @inicio AND H.DocDate < @fim
          AND H.CANCELED = 'N'
          AND L.BaseType = 202
        GROUP BY L.ItemCode
    """
    df = pd.read_sql(query, conn)
    df["MediaDiaria"] = df["Consumo3m"] / 90
    df["MediaMensal"] = df["Consumo3m"] / 3
    print(f"✅ Consumo MP calculado para {len(df)} itens")
    return df

def buscar_carteira_por_pa(conn):
    print("📋 Buscando carteira por PA...")
    query = """
        SELECT L.ItemCode, SUM(L.OpenQty) as Carteira
        FROM ORDR O WITH(NOLOCK)
        INNER JOIN RDR1 L WITH(NOLOCK) ON O.DocEntry=L.DocEntry
        WHERE O.DocStatus='O' AND L.OpenQty>0
        GROUP BY L.ItemCode
    """
    df = pd.read_sql(query, conn)
    print(f"✅ Carteira encontrada para {len(df)} PAs")
    return df

def buscar_historico_detalhado(conn):
    """NFs linha-a-linha dos últimos 3 meses fechados + mês atual.
    Aplica os mesmos filtros de buscar_consumo_3m:
      - NF não cancelada
      - VINCULADA A UM PEDIDO DE VENDA (BaseType=17 → ORDR)
      - NÃO é 1ª NF de Venda Futura (Comments NOT LIKE '%VD FUTURA%')"""
    print("📈 Buscando histórico detalhado (NF linha-a-linha)...")
    query = """
        DECLARE @inicio DATE = DATEADD(MONTH, -3, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
        DECLARE @fimMes DATE = DATEADD(MONTH, 1, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
        SELECT
            L.ItemCode,
            H.DocNum,
            O.DocNum AS Pedido,
            H.DocDate,
            H.CardCode,
            H.CardName,
            L.Quantity,
            L.LineTotal AS Valor
        FROM OINV H WITH(NOLOCK)
        INNER JOIN INV1 L WITH(NOLOCK) ON H.DocEntry = L.DocEntry
        INNER JOIN ORDR O WITH(NOLOCK) ON L.BaseEntry = O.DocEntry AND L.BaseType = 17
        WHERE H.DocDate >= @inicio AND H.DocDate < @fimMes
          AND H.CANCELED = 'N'
          AND ISNULL(H.Comments, '') NOT LIKE '%VD FUTURA%'
        ORDER BY L.ItemCode, H.DocDate DESC
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de NFs (com pedido) no histórico")
    return df

def buscar_historico_detalhado_mp(conn):
    """Goods Issues linha-a-linha de consumo de MP em produção (3 meses fechados + mês atual).
    Mesmos filtros de buscar_consumo_3m_mp: OIGE não cancelado + BaseType=202 (Production Order).
    LEFT JOIN com OWOR pra trazer DocNum da OP + ItemCode do PA produzido."""
    print("📈 Buscando histórico detalhado MP (Goods Issue linha-a-linha)...")
    query = """
        DECLARE @inicio DATE = DATEADD(MONTH, -3, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
        DECLARE @fimMes DATE = DATEADD(MONTH, 1, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
        SELECT
            L.ItemCode,
            H.DocNum    AS DocIssue,
            H.DocDate,
            W.DocNum    AS OP,
            W.ItemCode  AS PA_Produzido,
            PA.ItemName AS PA_Descricao,
            L.Quantity
        FROM OIGE H WITH(NOLOCK)
        INNER JOIN IGE1 L WITH(NOLOCK) ON H.DocEntry = L.DocEntry
        LEFT JOIN OWOR W WITH(NOLOCK) ON L.BaseEntry = W.DocEntry AND L.BaseType = 202
        LEFT JOIN OITM PA WITH(NOLOCK) ON W.ItemCode = PA.ItemCode
        WHERE H.DocDate >= @inicio AND H.DocDate < @fimMes
          AND H.CANCELED = 'N'
          AND L.BaseType = 202
        ORDER BY L.ItemCode, H.DocDate DESC
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de Goods Issue (vinculadas a OP) no histórico MP")
    return df

def buscar_todos_itens(conn):
    """Retorna ItemCode/ItemName/ItmsGrpCod/CountryOrg de TODOS os itens não cancelados.
    CountryOrg vira Fornecedor (BR=Nacional, outro=Importado). Usado pra resolver nomes/grupos
    na BOM e classificar fornecedor no Forecast Compra."""
    print("📋 Buscando descrições de todos os itens (incl. fantasmas)...")
    query = """
        SELECT ItemCode, ItemName, ItmsGrpCod, CountryOrg
        FROM OITM WITH(NOLOCK)
        WHERE Canceled = 'N'
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} itens cadastrados")
    return df

def buscar_compras_pa(conn):
    """Linhas individuais de pedidos de COMPRA de PAs (grupos 125 e 123) em aberto.
    Usada na drill-down do calendário de ruptura — seção 'PAs Chegando'."""
    print("📦 Buscando compras de PAs (linha por linha)...")
    query = """
        SELECT
            L.ItemCode,
            O.DocNum,
            O.DocDate,
            O.CardCode,
            O.CardName,
            L.OpenQty,
            L.ShipDate
        FROM OPOR O WITH(NOLOCK)
        INNER JOIN POR1 L WITH(NOLOCK) ON O.DocEntry=L.DocEntry
        INNER JOIN OITM T WITH(NOLOCK) ON L.ItemCode=T.ItemCode
        WHERE O.DocStatus='O' AND L.OpenQty>0
          AND T.ItmsGrpCod IN (125, 123)
        ORDER BY L.ItemCode, L.ShipDate
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de compra de PA")
    return df

def buscar_carteira_detalhada(conn):
    """Linhas individuais de pedidos de venda em aberto (ORDR/RDR1).
    Usada na drill-down do calendário de ruptura."""
    print("📋 Buscando carteira detalhada (linha por linha)...")
    query = """
        SELECT
            L.ItemCode,
            O.DocNum,
            O.DocDate,
            O.CardCode,
            O.CardName,
            L.OpenQty,
            L.ShipDate
        FROM ORDR O WITH(NOLOCK)
        INNER JOIN RDR1 L WITH(NOLOCK) ON O.DocEntry=L.DocEntry
        WHERE O.DocStatus='O' AND L.OpenQty>0
        ORDER BY L.ItemCode, L.ShipDate
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de pedidos em aberto")
    return df

def buscar_pedidos_compra(conn):
    print("🚚 Buscando pedidos de compra...")
    query = """
        SELECT L.ItemCode, SUM(L.OpenQty) as QtdChegando, MIN(L.ShipDate) as ProximaEntrega
        FROM OPOR O WITH(NOLOCK)
        INNER JOIN POR1 L WITH(NOLOCK) ON O.DocEntry=L.DocEntry
        WHERE O.DocStatus='O' AND L.OpenQty>0
        GROUP BY L.ItemCode
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} insumos com pedido em aberto")
    return df

def buscar_todos_pedidos_futuros(conn):
    print("📅 Buscando pedidos futuros...")
    query = """
        SELECT L.ItemCode, L.ShipDate, SUM(L.OpenQty) as QtdChegando
        FROM OPOR O WITH(NOLOCK)
        INNER JOIN POR1 L WITH(NOLOCK) ON O.DocEntry=L.DocEntry
        WHERE O.DocStatus='O' AND L.OpenQty>0 AND L.ShipDate>=GETDATE()
        GROUP BY L.ItemCode, L.ShipDate
    """
    return pd.read_sql(query, conn)

# ============================================================
# 🏷️  FAMÍLIA DE PRODUTOS
# ============================================================

def buscar_ordens_producao(conn):
    """Fase 2 — Ordens de produção abertas (OWOR)"""
    print("🏭 Buscando ordens de produção abertas...")
    query = """
        SELECT
            O.DocNum,
            O.ItemCode,
            T.ItemName,
            O.PlannedQty,
            ISNULL(O.CmpltQty, 0) as CmpltQty,
            O.PlannedQty - ISNULL(O.CmpltQty, 0) as SaldoOrdem,
            O.DueDate,
            O.Status
        FROM OWOR O WITH(NOLOCK)
        INNER JOIN OITM T WITH(NOLOCK) ON O.ItemCode = T.ItemCode
        WHERE O.Status IN ('R','P')
          AND O.PlannedQty > ISNULL(O.CmpltQty, 0)
        ORDER BY O.DueDate
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} ordens de produção abertas")
    return df

# ============================================================
# 🛒  COMPRAS — SC, Esboços, Workflow
# ============================================================

def buscar_sc_aberta(conn):
    """Cabeçalho das SCs (Solicitações de Compra) abertas — OPRQ DocStatus='O'.
    Junta com OUDP pra trazer nome do departamento (Department='-2' = sem depto)."""
    print("🛒 Buscando SCs abertas...")
    query = """
        SELECT
            H.DocEntry,
            H.DocNum,
            H.DocDate,
            H.DocDueDate,
            H.Requester,
            H.ReqName,
            H.Department      AS DepartmentCode,
            D.Name            AS DepartmentName,
            H.Comments,
            (SELECT COUNT(*) FROM PRQ1 L WITH(NOLOCK) WHERE L.DocEntry=H.DocEntry) AS QtdLinhas
        FROM OPRQ H WITH(NOLOCK)
        LEFT JOIN OUDP D WITH(NOLOCK) ON H.Department = D.Code
        WHERE H.DocStatus = 'O'
        ORDER BY H.DocDate DESC
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} SCs abertas")
    return df

def buscar_sc_linhas(conn):
    """Linhas das SCs abertas — item, qtd, observação, centro de custo."""
    print("🛒 Buscando linhas das SCs abertas...")
    query = """
        SELECT
            L.DocEntry,
            L.LineNum,
            L.ItemCode,
            L.Dscription   AS Descricao,
            L.Quantity,
            L.OcrCode      AS CentroCusto,
            L.Text         AS Observacao,
            L.ShipDate,
            L.LineStatus
        FROM PRQ1 L WITH(NOLOCK)
        INNER JOIN OPRQ H WITH(NOLOCK) ON H.DocEntry = L.DocEntry
        WHERE H.DocStatus = 'O'
        ORDER BY L.DocEntry, L.LineNum
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de SC abertas")
    return df

def buscar_esbocos_po(conn):
    """Esboços de PO (ODRF ObjType=22) ainda em aberto.
    UserSign → comprador via OUSR.U_NAME."""
    print("📝 Buscando esboços de PO abertos...")
    query = """
        SELECT
            D.DocEntry,
            D.DocNum,
            D.DocDate,
            D.CardCode,
            D.CardName,
            D.DocTotal,
            D.UserSign       AS CompradorId,
            U.U_NAME         AS Comprador,
            D.Comments
        FROM ODRF D WITH(NOLOCK)
        LEFT JOIN OUSR U WITH(NOLOCK) ON D.UserSign = U.USERID
        WHERE D.ObjType = 22
          AND D.DocStatus = 'O'
        ORDER BY D.DocDate DESC
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} esboços de PO abertos")
    return df

def buscar_esbocos_linhas(conn):
    """Linhas dos esboços de PO abertos (DRF1) — item, qtd, valor, obs, SC origem."""
    print("📝 Buscando linhas dos esboços de PO abertos...")
    query = """
        SELECT
            L.DocEntry,
            L.LineNum,
            L.ItemCode,
            L.Dscription   AS Descricao,
            L.Quantity,
            L.Price,
            L.LineTotal,
            L.ShipDate,
            L.Text         AS Observacao,
            L.BaseEntry    AS SC_DocEntry,
            L.BaseType
        FROM DRF1 L WITH(NOLOCK)
        INNER JOIN ODRF D WITH(NOLOCK) ON D.DocEntry = L.DocEntry
            AND D.ObjType = 22 AND D.DocStatus = 'O'
        ORDER BY L.DocEntry, L.LineNum
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de esboço")
    return df

def buscar_sc_esboco_link(conn):
    """Mapeamento SC ↔ Esboço de PO (relação N:M via DRF1.BaseEntry/BaseType=1470000113).
    Inclui apenas esboços abertos vinculados a SCs abertas."""
    print("🔗 Buscando vínculos SC ↔ Esboço...")
    query = """
        SELECT DISTINCT
            SC.DocEntry  AS SC_DocEntry,
            SC.DocNum    AS SC_DocNum,
            D.DocEntry   AS Esboco_DocEntry,
            D.DocNum     AS Esboco_DocNum
        FROM DRF1 L WITH(NOLOCK)
        INNER JOIN ODRF D WITH(NOLOCK) ON D.DocEntry = L.DocEntry
            AND D.ObjType = 22 AND D.DocStatus = 'O'
        INNER JOIN OPRQ SC WITH(NOLOCK) ON SC.DocEntry = L.BaseEntry
            AND L.BaseType = 1470000113 AND SC.DocStatus = 'O'
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} vínculos SC↔Esboço")
    return df

def buscar_aprovacoes_wf(conn):
    """Workflow de aprovação dos esboços de PO em workflow ATIVO (OWDD + WDD1 + OUSR).

    Vínculo correto: `OWDD.DraftEntry = ODRF.DocEntry` (em drafts, OWDD.DocEntry é null
    porque o documento final ainda não existe).

    Filtro `OWDD.ProcesStat='W'` é o que indica "workflow realmente em andamento" — bate
    com a tela "Documentos para Aprovação" do SAP B1. `OWDD.Status` sozinho pode ficar
    como 'W' fantasma mesmo depois do workflow terminar (lixo histórico).

    Status em WDD1 (por aprovador): W=aguardando, Y=aprovou, N=rejeitou."""
    print("✍️  Buscando workflow de aprovações ativas (ProcesStat=W)...")
    query = """
        SELECT
            W.WddCode,
            W.DraftEntry     AS Esboco_DocEntry,
            W.ObjType,
            W.DocDate        AS WorkflowDate,
            W.Status         AS WorkflowStatus,
            W.CurrStep       AS EtapaAtual,
            L.UserID         AS AprovadorId,
            U.U_NAME         AS Aprovador,
            L.Status         AS StatusAprovador,
            L.UpdateDate,
            L.Remarks,
            L.StepCode
        FROM OWDD W WITH(NOLOCK)
        INNER JOIN WDD1 L WITH(NOLOCK) ON W.WddCode = L.WddCode
        INNER JOIN ODRF D WITH(NOLOCK) ON D.DocEntry = W.DraftEntry
            AND D.ObjType = 22 AND D.DocStatus = 'O'
        LEFT JOIN OUSR U WITH(NOLOCK) ON L.UserID = U.USERID
        WHERE W.ObjType = 22
          AND W.ProcesStat = 'W'
        ORDER BY W.WddCode DESC, L.StepCode, L.UserID
    """
    df = pd.read_sql(query, conn)
    print(f"✅ {len(df)} linhas de aprovação (workflow)")
    return df


def calcular_mrp(df_pas_abc, pa_insumos, estoque_dict, pedidos_dict,
                 carteira_dict, nome_dict, df_ordens,
                 bom_dict=None, grupo_dict=None):
    """
    Fase 2 — MRP em cascata:
    - Explode necessidades brutas (carteira + ordens de produção abertas)
    - Desconta estoque disponível e pedidos em aberto
    - Cascata: PIs/CJs (grupo 124/126) com falta líquida propagam demanda pros
      seus filhos (componentes). Ex: PI 19.003.0009 com falta 50 → +50× cada
      componente da BOM (TINTA, TAMPA crua) na demanda dos filhos.
    - Coluna Tipo: 🏭 Fabricar (PI/CJ) ou 🛒 Comprar (insumos finais).
    """
    print("⚙️  Calculando MRP — necessidades de compra...")

    hoje = datetime.date.today()
    bom_dict   = bom_dict or {}
    grupo_dict = grupo_dict or {}

    # Saldo das ordens de produção abertas por PA
    ordens_dict = {}
    if len(df_ordens) > 0:
        for _, row in df_ordens.iterrows():
            pa    = row["ItemCode"]
            saldo = float(row.get("SaldoOrdem", 0) or 0)
            ordens_dict[pa] = ordens_dict.get(pa, 0) + saldo

    # Necessidade bruta por insumo
    nec_bruta  = {}   # ins -> qtd total necessária
    nec_por_pa = {}   # ins -> {pa: qtd} para rastreabilidade
    consumo_diario_por_insumo = {}  # ins -> consumo diário total

    for _, pa in df_pas_abc.iterrows():
        pa_code  = pa["ItemCode"]
        ins_dict = pa_insumos.get(pa_code, {})
        if not ins_dict:
            continue

        carteira     = carteira_dict.get(pa_code, 0)
        ordens       = ordens_dict.get(pa_code, 0)
        estoque_pa   = float(pa.get("Estoque", 0) or 0)
        # Demanda = máximo entre carteira e ordens (evita dupla contagem entre PV e OP).
        # Necessidade real = demanda menos o estoque do PA já pronto (cobre carteira sem produzir).
        demanda_pa = max(carteira, ordens)
        nec_pa     = max(0, demanda_pa - estoque_pa)

        if nec_pa <= 0:
            continue

        media_dia = pa["MediaDiaria"]

        for ins, qtd_por_pa in ins_dict.items():
            if qtd_por_pa <= 0:
                continue
            nec = nec_pa * qtd_por_pa
            nec_bruta[ins] = nec_bruta.get(ins, 0) + nec
            nec_por_pa.setdefault(ins, {})[pa_code] = nec
            consumo_diario_por_insumo[ins] = (
                consumo_diario_por_insumo.get(ins, 0) + media_dia * qtd_por_pa
            )

    # ── CASCATA: PIs/CJs com falta líquida propagam demanda pros filhos ──
    # Itera várias vezes pra cobrir cascateamentos profundos (PI → CJ → componentes).
    propagado = {}   # tracking por (pai, filho) pra evitar dupla contagem na re-iteração
    for _ in range(8):
        mudou = False
        for ins in list(nec_bruta.keys()):
            grupo = grupo_dict.get(ins, 0)
            if grupo not in GRUPOS_RECURSIVOS:    # só PI (124) e CJ (126)
                continue
            if ins not in bom_dict:
                continue
            estoque  = estoque_dict.get(ins, 0)
            ped      = pedidos_dict.get(ins, {})
            chegando = float(ped.get("QtdChegando", 0) or 0)
            falta    = max(0, nec_bruta[ins] - estoque - chegando)
            ja_propagado = propagado.get(ins, 0)
            delta_falta  = falta - ja_propagado
            if delta_falta <= 0:
                continue
            # Propaga delta da falta pros filhos via BOM
            for (filho, qtd_filho) in bom_dict[ins]:
                grupo_filho = grupo_dict.get(filho, 0)
                if grupo_filho in GRUPOS_IGNORAR:
                    continue
                acrescimo = delta_falta * qtd_filho
                nec_bruta[filho] = nec_bruta.get(filho, 0) + acrescimo
                # Rastreio de origem (qual PI demandou)
                nec_por_pa.setdefault(filho, {})[f"<via {ins}>"] = (
                    nec_por_pa.get(filho, {}).get(f"<via {ins}>", 0) + acrescimo
                )
                # Propaga consumo diário proporcional
                consumo_pi = consumo_diario_por_insumo.get(ins, 0)
                if consumo_pi > 0:
                    consumo_diario_por_insumo[filho] = (
                        consumo_diario_por_insumo.get(filho, 0) + consumo_pi * qtd_filho
                    )
            propagado[ins] = falta
            mudou = True
        if not mudou:
            break

    # Necessidade líquida = bruta - estoque - pedidos em aberto - OPs produzindo o próprio item.
    # OPs com ItemCode == insumo (caso PI/CJ) representam supply: vão produzir o item.
    # Incluímos TODOS os itens com demanda (mesmo nec_liquida=0) — a página MRP filtra
    # por "Só com falta" por padrão mas o user pode ver "Todos os insumos demandados".
    sugestoes = []
    for ins, nec_b in nec_bruta.items():
        estoque_atual = estoque_dict.get(ins, 0)
        ped           = pedidos_dict.get(ins, {})
        qtd_chegando  = float(ped.get("QtdChegando", 0) or 0)
        prox_entrega  = ped.get("ProximaEntrega", None)
        op_produzindo = float(ordens_dict.get(ins, 0) or 0)   # OPs produzindo este item (PIs)

        nec_liquida = max(0, nec_b - estoque_atual - qtd_chegando - op_produzindo)

        # Data de necessidade — quando o estoque vai zerar
        consumo_dia = consumo_diario_por_insumo.get(ins, 0)
        if nec_liquida <= 0:
            data_necessidade = None   # supply cobre, sem prazo
        elif consumo_dia > 0 and estoque_atual > 0:
            dias_cobertura   = estoque_atual / consumo_dia
            data_necessidade = hoje + datetime.timedelta(days=int(dias_cobertura))
        else:
            data_necessidade = hoje  # já está em falta

        # Urgência
        if nec_liquida <= 0:
            urgencia = "🟢 OK"
        else:
            dias_para_nec = (data_necessidade - hoje).days
            if dias_para_nec <= 0:
                urgencia = "🔴 URGENTE"
            elif dias_para_nec <= 7:
                urgencia = "🟡 ATENÇÃO"
            else:
                urgencia = "🟢 OK"

        # Tipo: Fabricar (PI/CJ) ou Comprar (demais)
        grupo_ins = grupo_dict.get(ins, 0)
        tipo = "🏭 Fabricar" if grupo_ins in GRUPOS_RECURSIVOS else "🛒 Comprar"

        # PAs afetados (inclui marcadores "<via ...>")
        pas_dict    = nec_por_pa.get(ins, {})
        pas_lista   = sorted(pas_dict.keys(), key=lambda p: pas_dict[p], reverse=True)
        pas_str     = ", ".join(pas_lista)

        sugestoes.append({
            "Urgência":          urgencia,
            "Tipo":              tipo,
            "Insumo":            ins,
            "Descrição":         str(nome_dict.get(ins, "")).strip(),
            "Estoque Atual":     int(estoque_atual),
            "Pedido em Aberto":  "SIM" if qtd_chegando > 0 else "—",
            "Qtd Chegando":      int(qtd_chegando),
            "OP Produzindo":     int(op_produzindo),
            "Prev. Chegada":     prox_entrega.strftime("%d/%m/%Y") if pd.notna(prox_entrega) and prox_entrega else "—",
            "Nec. Bruta":        int(nec_b),
            "Nec. Líquida":      int(nec_liquida),
            "Lead Time (dias)":  "A definir",
            "Data Necessidade":  data_necessidade.strftime("%d/%m/%Y") if data_necessidade else "—",
            "Emitir Pedido Até": "A definir (aguarda lead time)" if nec_liquida > 0 else "—",
            "PAs Afetados":      pas_str,
        })

    df_mrp = pd.DataFrame(sugestoes) if sugestoes else pd.DataFrame()
    if len(df_mrp) > 0:
        urgencia_ord = {"🔴 URGENTE": 0, "🟡 ATENÇÃO": 1, "🟢 OK": 2}
        df_mrp["_ord"] = df_mrp["Urgência"].map(urgencia_ord).fillna(3)
        df_mrp = df_mrp.sort_values(["_ord", "Data Necessidade"]).drop(columns=["_ord"])

    print(f"✅ MRP calculado: {len(df_mrp)} insumos precisam de reposição")
    return df_mrp


def _carregar_regras_target(path=os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados", "targets_estoque.csv")):
    """Carrega regras de target_meses por padrão de código. Retorna (rules, default)."""
    import os as _os
    if not _os.path.exists(path):
        return [], 3.0
    try:
        df = pd.read_csv(path, comment="#")
    except Exception:
        return [], 3.0
    rules, default = [], 3.0
    for _, r in df.iterrows():
        tipo = str(r.get("regra_tipo","")).strip().lower()
        pad  = str(r.get("padrao","")).strip()
        try: meses = float(r.get("target_meses", 3))
        except Exception: meses = 3.0
        if tipo == "default":
            default = meses
        elif tipo in ("prefix","exact","contains") and pad:
            rules.append((tipo, pad.upper(), meses))
    # Ordena por especificidade (padrão mais longo primeiro)
    rules.sort(key=lambda x: -len(x[1]))
    return rules, default


def _get_target_meses(codigo, rules, default):
    cod = str(codigo or "").upper()
    for (tipo, pad, meses) in rules:
        if tipo == "prefix"   and cod.startswith(pad): return meses
        if tipo == "exact"    and cod == pad:           return meses
        if tipo == "contains" and pad in cod:           return meses
    return default


def _carregar_overrides_target(path=os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados", "target_overrides.csv")):
    """Overrides editados pelo PCP via UI. Item-específico (ItemCode → meses)."""
    import os as _os
    if not _os.path.exists(path):
        return {}
    try:
        df = pd.read_csv(path)
        return {str(r["Código"]).strip(): float(r["Target Meses"]) for _, r in df.iterrows()
                if pd.notna(r.get("Código")) and pd.notna(r.get("Target Meses"))}
    except Exception:
        return {}


def calcular_forecast_compra(df_pas_abc, df_bom, df_estoque_insumos, df_pedidos,
                              nome_dict_full, grupo_dict_full,
                              fonte_dict=None):
    """Forecast de Compra inspirado na planilha 'Forecast Mensal China'.

    Pra cada insumo:
    - **Forecast (4 meses)** = soma da demanda dos próximos 4 meses, calculada como
      Média Mensal (3 meses fechados) × Qtd_p_1_PA × 4. Demanda agregada de TODOS PAs
      que usam o insumo, separada em colunas por modelo (LIH, LHPG2, LPP, LNP, EVO,
      HBM, Revenda).
    - **Estoque Atual** vem de OITW (mesmos armazéns).
    - **Tempo de Estoque (meses)** = Estoque / Média Mensal.
    - **Em Trânsito + Pedido** = saldo de POs em aberto.
    - **Target Estoque (3 meses)** = Média × 3 (default).
    - **Compra para Repor** = max(0, Target − Estoque − Em Trânsito).

    Pra PAs de Revenda (grupo 123), o próprio PA aparece como "insumo a comprar"
    (já que não tem BOM — é comprado pronto).
    """
    print("📦 Calculando Forecast de Compra...")

    target_rules, target_default = _carregar_regras_target()
    overrides = _carregar_overrides_target()
    print(f"   {len(target_rules)} regras de target + {len(overrides)} overrides (default {target_default} meses)")
    fonte_dict = fonte_dict or {}

    MODELOS_ORD = ["LIH","LHPG2","LPP","LNP","EVO","HBM","Revenda"]
    MODELO_MAP = {
        "LIH": "LIH",
        "LHP G2": "LHPG2", "LHPG2": "LHPG2",
        "LPP": "LPP",
        "LNP": "LNP",
        "EVO": "EVO",
        "HBM": "HBM",
    }

    # bom_dict
    bom_dict = {}
    for _, r in df_bom.iterrows():
        bom_dict.setdefault(r["Father"], []).append((r["Insumo"], float(r["Quantity"])))

    estoque_dict = df_estoque_insumos.set_index("ItemCode")["Estoque"].to_dict()
    pedidos_dict = df_pedidos.set_index("ItemCode").to_dict("index") if len(df_pedidos) > 0 else {}

    # Mapeia PA → modelo + forecast (4 meses)
    pa_modelo, pa_forecast = {}, {}
    for _, pa in df_pas_abc.iterrows():
        pa_code = pa["ItemCode"]
        familia = pa.get("Familia", "") or ""
        if pa.get("ItmsGrpCod") == 123:
            modelo = "Revenda"
        else:
            modelo = MODELO_MAP.get(familia, "Outros")
        media_mensal_pa = float(pa.get("MediaMensal", 0) or 0)
        pa_modelo[pa_code] = modelo
        pa_forecast[pa_code] = media_mensal_pa * 4   # próximos 4 meses

    # Acumula demanda por insumo, separada por modelo.
    # Pra grupo 125 (PA fabricado), expandir_bom sem netting (estoque vazio) → demanda bruta.
    # Pra grupo 123 (Revenda), o próprio PA entra como "insumo" da coluna Revenda.
    forecast_por_insumo = {}   # ins → {modelo: qtd_total_4_meses}

    for pa_code, forecast_pa in pa_forecast.items():
        if forecast_pa <= 0:
            continue
        modelo = pa_modelo[pa_code]

        if modelo == "Revenda":
            forecast_por_insumo.setdefault(pa_code, {})
            forecast_por_insumo[pa_code][modelo] = (
                forecast_por_insumo[pa_code].get(modelo, 0) + forecast_pa
            )
            continue

        # PA fabricado — expande BOM com estoque vazio (forecast = demanda bruta)
        if pa_code not in bom_dict:
            continue
        # expandir_bom retorna lista de (componente, qtd) — usamos {} como estoque pra
        # forçar explosão completa até insumos finais (sem netting, sem stop em PI/CJ)
        for (comp, qtd_comp) in bom_dict[pa_code]:
            comps = expandir_bom(comp, qtd_comp * forecast_pa, bom_dict, {}, grupo_dict_full)
            for (ins, q) in comps:
                forecast_por_insumo.setdefault(ins, {})
                forecast_por_insumo[ins][modelo] = (
                    forecast_por_insumo[ins].get(modelo, 0) + q
                )

    # Monta DataFrame de saída
    resultados = []
    for ins, modelos in forecast_por_insumo.items():
        total = sum(modelos.values())
        if total <= 0:
            continue

        nome      = str(nome_dict_full.get(ins, "")).strip()
        estoque   = float(estoque_dict.get(ins, 0) or 0)
        ped       = pedidos_dict.get(ins, {}) if isinstance(pedidos_dict.get(ins), dict) else {}
        em_trans  = float(ped.get("QtdChegando", 0) or 0)
        media_mes = total / 4.0
        tempo_est = round(estoque / media_mes, 2) if media_mes > 0 else 0
        # Target específico por item: override (UI) > regras CSV > default
        if ins in overrides:
            target_meses = float(overrides[ins])
        else:
            target_meses = _get_target_meses(ins, target_rules, target_default)
        target       = media_mes * target_meses
        compra       = max(0.0, target - estoque - em_trans)

        # Fornecedor: BR=Nacional, outro=Importado, vazio=inferir por prefixo
        country = str(fonte_dict.get(ins, "") or "").strip().upper()
        if country == "BR":
            fornecedor = "Nacional"
        elif country in ("CN","US","DE","JP","KR","TW","IN","MX","IT","FR","UK","GB","ES"):
            fornecedor = "Importado"
        else:
            # Heuristic: codes with prefix IN... are typically imported items
            fornecedor = "Importado" if str(ins).upper().startswith("IN") else "Nacional"

        linha = {
            "Código":               ins,
            "Descrição":            nome,
            "Fornecedor":           fornecedor,
            "Forecast (4 meses)":   int(round(total)),
            "Média Mensal":         int(round(media_mes)),
        }
        for m in MODELOS_ORD:
            linha[m] = int(round(modelos.get(m, 0)))
        linha.update({
            "Estoque Atual":            int(round(estoque)),
            "Tempo de Estoque (meses)": tempo_est,
            "Em Trânsito + Pedido":     int(round(em_trans)),
            "Target Meses":             target_meses,
            "Target Estoque":           int(round(target)),
            "Compra para Repor":        int(round(compra)),
        })
        resultados.append(linha)

    df = pd.DataFrame(resultados)
    if len(df) > 0:
        df = df.sort_values("Compra para Repor", ascending=False).reset_index(drop=True)
    print(f"✅ Forecast Compra: {len(df)} itens")
    return df


def calcular_mrp_revenda(df_pas_abc, pedidos_dict, carteira_dict):
    """
    MRP para produtos de REVENDA (ItmsGrpCod=123) — não são produzidos, são comprados.
    Necessidade = max(carteira, média diária × horizonte) - estoque - pedidos em aberto.
    """
    print("⚙️  Calculando MRP Revenda...")
    hoje = datetime.date.today()
    HORIZONTE_DIAS = 60  # olha 60 dias à frente

    sugestoes = []
    df_rev = df_pas_abc[df_pas_abc["ItmsGrpCod"] == 123]

    for _, pa in df_rev.iterrows():
        pa_code   = pa["ItemCode"]
        estoque   = int(pa.get("Estoque", 0) or 0)
        media_dia = float(pa.get("MediaDiaria", 0) or 0)
        carteira  = int(carteira_dict.get(pa_code, 0) or 0)

        ped          = pedidos_dict.get(pa_code, {})
        qtd_chegando = float(ped.get("QtdChegando", 0) or 0)
        prox_entrega = ped.get("ProximaEntrega", None)

        nec_bruta   = max(carteira, int(media_dia * HORIZONTE_DIAS))
        nec_liquida = max(0, nec_bruta - estoque - int(qtd_chegando))
        if nec_liquida <= 0:
            continue

        if media_dia > 0 and estoque > 0:
            dias_cobertura   = estoque / media_dia
            data_necessidade = hoje + datetime.timedelta(days=int(dias_cobertura))
        else:
            data_necessidade = hoje

        dias_para_nec = (data_necessidade - hoje).days
        if dias_para_nec <= 0:
            urgencia = "🔴 URGENTE"
        elif dias_para_nec <= 7:
            urgencia = "🟡 ATENÇÃO"
        else:
            urgencia = "🟢 OK"

        media_mes = media_dia * 30
        sugestoes.append({
            "Urgência":          urgencia,
            "Família":           pa.get("Familia", ""),
            "Código PA":         pa_code,
            "Descrição":         str(pa.get("ItemName", "")).strip(),
            "Classe ABC":        pa.get("ABC", ""),
            "Estoque Atual":     estoque,
            "Carteira":          carteira,
            "Média/mês":         int(round(media_mes)),
            "Cob. (meses)":      round(estoque / media_mes, 1) if media_mes > 0 else "-",
            "Pedido em Aberto":  "SIM" if qtd_chegando > 0 else "—",
            "Qtd Chegando":      int(qtd_chegando),
            "Prev. Chegada":     prox_entrega.strftime("%d/%m/%Y") if pd.notna(prox_entrega) and prox_entrega else "—",
            "Nec. Bruta":        int(nec_bruta),
            "Nec. Líquida":      int(nec_liquida),
            "Data Necessidade":  data_necessidade.strftime("%d/%m/%Y"),
        })

    df_rev_mrp = pd.DataFrame(sugestoes) if sugestoes else pd.DataFrame()
    if len(df_rev_mrp) > 0:
        urg_ord = {"🔴 URGENTE": 0, "🟡 ATENÇÃO": 1, "🟢 OK": 2}
        df_rev_mrp["_ord"] = df_rev_mrp["Urgência"].map(urg_ord).fillna(3)
        df_rev_mrp = df_rev_mrp.sort_values(["_ord", "Data Necessidade"]).drop(columns=["_ord"])

    print(f"✅ MRP Revenda: {len(df_rev_mrp)} produtos precisam de reposição")
    return df_rev_mrp


def extrair_familia(codigo):
    try:
        cod = codigo.replace("PREFIX", "")
        modelo = re.match(r'^([A-Z]+)', cod)
        if not modelo:
            return "OUTROS"
        modelo = modelo.group(1)
        if codigo.endswith("G2"):
            return f"{modelo} G2"
        return modelo
    except:
        return "OUTROS"

# ============================================================
# 🧮  CURVA ABC
# ============================================================

def calcular_curva_abc(df_pas, df_consumo):
    print("📈 Calculando curva ABC...")
    df = df_pas.merge(df_consumo[["ItemCode","Consumo3m","MediaDiaria","MediaMensal"]], on="ItemCode", how="left")
    df["MediaMensal"] = df["MediaMensal"].fillna(0)
    df["Consumo3m"]   = df["Consumo3m"].fillna(0)
    df["MediaDiaria"] = df["MediaDiaria"].fillna(0)
    df["Familia"]     = df["ItemCode"].apply(extrair_familia)
    df = df.sort_values("Consumo3m", ascending=False)
    total = df["Consumo3m"].sum()
    if total > 0:
        df["PctAcum"] = df["Consumo3m"].cumsum() / total * 100
        df["ABC"]     = df["PctAcum"].apply(lambda x: "A" if x<=80 else ("B" if x<=95 else "C"))
    else:
        df["PctAcum"] = 0
        df["ABC"]     = "C"
    return df

# ============================================================
# 🔄  EXPANDIR BOM MULTINÍVEL
# ============================================================

def expandir_bom(item_code, qtd_necessaria, bom_dict, estoque_dict, grupo_dict, nivel=0):
    """Expansão multinível com **netting** em cada nível.

    Pra cada item intermediário com BOM, considera o estoque dele:
    - Se estoque cobre tudo → para, não desce.
    - Se estoque cobre parte → desce só com o saldo (qtd_necessaria - estoque).
    - Se não tem estoque → desce com a qtd inteira.

    Exemplo: precisa 100 TAMPA pintada (19.006.0059), tem 30 em estoque.
    Desce pedindo TAMPA sem pintar + TINTA pra **70** (não pra 100). Os 30
    pintados já cobrem parte da demanda.
    """
    if nivel > 6:
        return [(item_code, qtd_necessaria)]

    grupo = grupo_dict.get(item_code, 0)

    if grupo in GRUPOS_IGNORAR:
        return []

    tem_bom      = item_code in bom_dict
    estoque      = estoque_dict.get(item_code, 0)
    eh_recursivo = grupo in GRUPOS_RECURSIVOS

    forca_descer = (
        tem_bom
        and grupo == 120
        and isinstance(item_code, str)
        and item_code.startswith("F07")
    )

    if not tem_bom:
        return [(item_code, qtd_necessaria)]

    # Netting: estoque do intermediário cobre parte (ou toda) a demanda.
    # forca_descer e nivel==0 (PA raiz) nunca consomem do próprio estoque (precisa explodir tudo).
    if not forca_descer and (eh_recursivo or nivel > 0):
        coberto = min(estoque, qtd_necessaria) if estoque > 0 else 0
        falta   = qtd_necessaria - coberto
    else:
        coberto = 0
        falta   = qtd_necessaria

    # Estoque cobriu tudo — não precisa explodir
    if falta <= 0:
        return [(item_code, qtd_necessaria)]

    # Explode SÓ a parte que falta produzir/comprar
    resultado = []
    for (comp, qtd_comp) in bom_dict[item_code]:
        sub = expandir_bom(comp, qtd_comp * falta, bom_dict, estoque_dict, grupo_dict, nivel+1)
        resultado.extend(sub)
    return resultado

# ============================================================
# 🔢  CALCULAR CAPACIDADE — CASCATA EM PRIORIDADE
#
# Lógica:
# 1. Ordena PAs por prioridade (ABC + carteira > estoque + média/dia)
# 2. Para cada PA na ordem, calcula quantas peças consegue montar
#    com o estoque RESTANTE dos insumos (após consumo dos PAs anteriores)
# 3. Consome os insumos proporcionalmente
# 4. Resultado: cada PA vê o estoque real disponível após os anteriores
# ============================================================

def calcular_capacidade(df_pas_abc, df_bom, df_estoque_insumos, df_pedidos, df_carteira):
    print("🔢 Calculando capacidade de produção (cascata em prioridade)...")

    estoque_dict  = df_estoque_insumos.set_index("ItemCode")["Estoque"].to_dict()
    nome_dict     = df_estoque_insumos.set_index("ItemCode")["ItemName"].to_dict()
    grupo_dict    = df_estoque_insumos.set_index("ItemCode")["ItmsGrpCod"].to_dict()
    pedidos_dict  = df_pedidos.set_index("ItemCode").to_dict("index") if len(df_pedidos) > 0 else {}
    carteira_dict = df_carteira.set_index("ItemCode")["Carteira"].to_dict() if len(df_carteira) > 0 else {}

    bom_dict = {}
    for _, row in df_bom.iterrows():
        f = row["Father"]
        if f not in bom_dict:
            bom_dict[f] = []
        bom_dict[f].append((row["Insumo"], row["Quantity"]))

    # Estoque disponível = estoque atual + pedidos em aberto
    estoque_disponivel = dict(estoque_dict)
    for ins, ped in pedidos_dict.items():
        qtd_chegando = ped.get("QtdChegando", 0) or 0
        estoque_disponivel[ins] = estoque_disponivel.get(ins, 0) + qtd_chegando

    # Expande BOM de todos os PAs
    pa_insumos = {}
    for _, pa in df_pas_abc.iterrows():
        pa_code = pa["ItemCode"]
        componentes = bom_dict.get(pa_code, [])
        ins_dict = {}
        for (comp, qtd) in componentes:
            exp = expandir_bom(comp, qtd, bom_dict, estoque_dict, grupo_dict)
            for (ins, q) in exp:
                ins_dict[ins] = ins_dict.get(ins, 0) + q
        pa_insumos[pa_code] = ins_dict

    # Calcula score de prioridade de cada PA
    # Critérios: ABC (A=3, B=2, C=1) + urgência (carteira > estoque) + média/dia + carteira
    def score_pa(pa):
        abc_score = {"A": 3, "B": 2, "C": 1}.get(pa["ABC"], 0)
        media     = pa["MediaDiaria"]
        carteira  = carteira_dict.get(pa["ItemCode"], 0)
        estoque   = pa["Estoque"]
        urgente   = 2.0 if carteira > estoque else 1.0
        return (abc_score * urgente * 1000) + (media * 10) + carteira

    # Ordena PAs por prioridade (maior score primeiro)
    df_ordenado = df_pas_abc.copy()
    df_ordenado["_score"] = df_ordenado.apply(score_pa, axis=1)
    df_ordenado = df_ordenado.sort_values("_score", ascending=False)

    # Estoque disponível flutuante — vai sendo consumido em cascata
    estoque_restante = dict(estoque_disponivel)

    resultados = []
    for _, pa in df_ordenado.iterrows():
        pa_code   = pa["ItemCode"]
        ins_dict  = pa_insumos.get(pa_code, {})
        media_dia = pa["MediaDiaria"]
        carteira  = carteira_dict.get(pa_code, 0)

        eh_revenda = pa.get("ItmsGrpCod", 0) == 123
        media_mes = media_dia * 30
        if eh_revenda or not ins_dict:
            # Produto de revenda — não é produzido, capacidade = estoque + pedidos em aberto
            ped_pa       = pedidos_dict.get(pa_code, {}) or {}
            qtd_cheg_pa  = float(ped_pa.get("QtdChegando", 0) or 0)
            prox_entr_pa = ped_pa.get("ProximaEntrega", None)
            estoque_pa   = int(pa["Estoque"])
            capacidade   = estoque_pa + int(qtd_cheg_pa)
            resultados.append({
                "Família":            pa["Familia"],
                "Categoria":          "Revenda" if eh_revenda else "PA",
                "Código PA":          pa_code,
                "Descrição PA":       pa["ItemName"],
                "Classe ABC":         pa["ABC"],
                "Estoque PA":         estoque_pa,
                "Carteira":           int(carteira),
                "Média/mês":          int(round(media_mes)),
                "Cob. PA (meses)":    round(pa["Estoque"] / media_mes, 1) if media_mes > 0 else "-",
                "Insumo Gargalo":     "REVENDA",
                "Estoque Insumo":     capacidade,
                "Peças Possíveis":    capacidade,
                "Cob. Insumo (meses)":round(capacidade / media_mes, 1) if media_mes > 0 else "-",
                "Próx. Entrega":      prox_entr_pa,
                "Qtd Chegando":       int(qtd_cheg_pa),
                "Insumo Zerado":      "Revenda — não produzido",
            })
            continue

        # Calcula quantas peças consegue montar com o estoque RESTANTE
        min_pecas       = float("inf")
        insumo_gargalo  = ""
        estoque_gargalo = 0

        for ins, qtd_necessaria in ins_dict.items():
            if qtd_necessaria <= 0:
                continue
            est_disp = estoque_restante.get(ins, 0)
            pecas    = est_disp / qtd_necessaria
            if pecas < min_pecas:
                min_pecas       = pecas
                insumo_gargalo  = ins
                estoque_gargalo = estoque_dict.get(ins, 0)  # estoque original (sem pedidos)

        pecas_possiveis = int(min_pecas) if min_pecas != float("inf") else 0

        # Consome os insumos do estoque restante
        if pecas_possiveis > 0:
            for ins, qtd_necessaria in ins_dict.items():
                consumo_total = pecas_possiveis * qtd_necessaria
                estoque_restante[ins] = max(0, estoque_restante.get(ins, 0) - consumo_total)

        # Identifica insumo zerado (que causou o limite)
        insumo_zerado_txt = ""
        if pecas_possiveis == 0:
            # Mostra o insumo com menor estoque restante
            candidatos = [
                (estoque_restante.get(ins, 0), ins)
                for ins, qtd in ins_dict.items()
                if qtd > 0 and estoque_restante.get(ins, 0) < qtd
            ]
            if candidatos:
                candidatos.sort()
                _, ins_zerado = candidatos[0]
                nome_ins = str(nome_dict.get(ins_zerado, "")).strip()[:35]
                est_rest = estoque_restante.get(ins_zerado, 0)
                insumo_zerado_txt = f"{ins_zerado} - {nome_ins} (restam: {int(est_rest)})"

        # Pedido do insumo gargalo
        pedido    = pedidos_dict.get(insumo_gargalo, {})
        prox_entr = pedido.get("ProximaEntrega", None)
        qtd_cheg  = pedido.get("QtdChegando", 0)

        media_mes_pa = media_dia * 30
        cob_pa  = round(pa["Estoque"] / media_mes_pa, 1) if media_mes_pa > 0 else "-"
        cob_ins = round(pecas_possiveis / media_mes_pa, 1) if media_mes_pa > 0 and pecas_possiveis > 0 else 0

        resultados.append({
            "Família":            pa["Familia"],
            "Categoria":          "Revenda" if pa.get("ItmsGrpCod", 0) == 123 else "PA",
            "Código PA":          pa_code,
            "Descrição PA":       pa["ItemName"],
            "Classe ABC":         pa["ABC"],
            "Estoque PA":         int(pa["Estoque"]),
            "Carteira":           int(carteira),
            "Média/mês":          int(round(media_mes_pa)),
            "Cob. PA (meses)":    cob_pa,
            "Insumo Gargalo":     insumo_gargalo,
            "Estoque Insumo":     int(estoque_gargalo),
            "Peças Possíveis":    pecas_possiveis,
            "Cob. Insumo (meses)":cob_ins,
            "Próx. Entrega":      prox_entr,
            "Qtd Chegando":       int(qtd_cheg) if qtd_cheg else 0,
            "Insumo Zerado":      insumo_zerado_txt,
        })

    df_result = pd.DataFrame(resultados)
    # Reordena por família e ABC para exibição
    df_result = df_result.sort_values(["Família", "Classe ABC", "Cob. Insumo (meses)"])
    print(f"✅ Capacidade calculada para {len(df_result)} PAs")
    return df_result, pa_insumos, estoque_dict, pedidos_dict, carteira_dict, nome_dict

# ============================================================
# 📅  CALENDÁRIO DE RUPTURA
# ============================================================

def calcular_calendario(df_pas_abc, df_bom, df_estoque_insumos, df_pedidos_futuros,
                        df_carteira=None, df_compras_pa=None, df_ordens=None,
                        df_carteira_det=None):
    print("📅 Calculando calendário...")

    estoque_dict = df_estoque_insumos.set_index("ItemCode")["Estoque"].to_dict()
    grupo_dict   = df_estoque_insumos.set_index("ItemCode")["ItmsGrpCod"].to_dict()

    bom_dict = {}
    for _, row in df_bom.iterrows():
        f = row["Father"]
        if f not in bom_dict:
            bom_dict[f] = []
        bom_dict[f].append((row["Insumo"], row["Quantity"]))

    hoje = datetime.date.today()
    seg  = hoje - datetime.timedelta(days=hoje.weekday())
    semanas = [seg + datetime.timedelta(weeks=i) for i in range(52)]

    # Datas atrasadas (< hoje) ou nulas caem na semana 0.
    def _semana_idx(data_obj):
        if data_obj is None or pd.isna(data_obj):
            return 0
        d = data_obj.date() if hasattr(data_obj, "date") else data_obj
        if d < seg:
            return 0
        idx = (d - seg).days // 7
        return idx if idx < len(semanas) else None

    # Chegadas (POs + OPs) por PA por semana
    chegadas = {}  # {pa_code: {semana_idx: qtd}}

    if df_compras_pa is not None and len(df_compras_pa) > 0:
        _df = df_compras_pa.copy()
        _df["ShipDate"] = pd.to_datetime(_df["ShipDate"], errors="coerce")
        for _, row in _df.iterrows():
            pa_code = row["ItemCode"]
            qtd     = float(row.get("OpenQty", 0) or 0)
            if qtd <= 0:
                continue
            idx = _semana_idx(row["ShipDate"])
            if idx is None:
                continue
            chegadas.setdefault(pa_code, {})
            chegadas[pa_code][idx] = chegadas[pa_code].get(idx, 0) + qtd

    if df_ordens is not None and len(df_ordens) > 0:
        _df = df_ordens.copy()
        _df["DueDate"] = pd.to_datetime(_df["DueDate"], errors="coerce")
        for _, row in _df.iterrows():
            pa_code = row["ItemCode"]
            qtd     = float(row.get("SaldoOrdem", 0) or 0)
            if qtd <= 0:
                continue
            idx = _semana_idx(row["DueDate"])
            if idx is None:
                continue
            chegadas.setdefault(pa_code, {})
            chegadas[pa_code][idx] = chegadas[pa_code].get(idx, 0) + qtd

    # Vendas (carteira) por PA por semana — usa ShipDate de cada pedido individual.
    # Pedidos sem data ou com ShipDate vencida caem na semana 0.
    vendas_por_pa = {}  # {pa_code: {semana_idx: qtd}}
    if df_carteira_det is not None and len(df_carteira_det) > 0:
        _df = df_carteira_det.copy()
        _df["ShipDate"] = pd.to_datetime(_df["ShipDate"], errors="coerce")
        for _, row in _df.iterrows():
            pa_code = row["ItemCode"]
            qtd     = float(row.get("OpenQty", 0) or 0)
            if qtd <= 0:
                continue
            idx = _semana_idx(row["ShipDate"])
            if idx is None:
                continue
            vendas_por_pa.setdefault(pa_code, {})
            vendas_por_pa[pa_code][idx] = vendas_por_pa[pa_code].get(idx, 0) + qtd
    elif df_carteira is not None and len(df_carteira) > 0:
        # Fallback: sem detalhe por linha, abate carteira total na semana 0
        for _, row in df_carteira.iterrows():
            pa_code = row["ItemCode"]
            qtd     = float(row.get("Carteira", 0) or 0)
            if qtd <= 0:
                continue
            vendas_por_pa.setdefault(pa_code, {})
            vendas_por_pa[pa_code][0] = vendas_por_pa[pa_code].get(0, 0) + qtd

    calendario = []
    for _, pa in df_pas_abc.iterrows():
        pa_code   = pa["ItemCode"]
        media_dia = pa["MediaDiaria"]

        categoria = "Revenda" if pa.get("ItmsGrpCod", 0) == 123 else "PA"

        chegadas_pa_pre = chegadas.get(pa_code, {})
        vendas_pa_pre   = vendas_por_pa.get(pa_code, {})

        # PA sem histórico de consumo E sem carteira E sem chegadas → SEM HISTORICO em todas as semanas.
        # Mas se houver carteira ou chegada, processa normalmente — PCP precisa enxergar essa ruptura
        # mesmo sem média histórica (PA novo com pedido já confirmado, por exemplo).
        if media_dia <= 0 and not vendas_pa_pre and not chegadas_pa_pre:
            linha = {
                "Família":   pa["Familia"],
                "Categoria": categoria,
                "Código PA": pa_code,
                "Descrição": pa["ItemName"][:45],
                "ABC":       pa["ABC"],
            }
            for semana in semanas:
                label = f"S{semana.strftime('%V')}\n{semana.strftime('%d/%m')}"
                linha[label] = "SEM HISTORICO"
            calendario.append(linha)
            continue

        # Produtos de revenda não têm BOM — são incluídos só com base no estoque
        linha = {
            "Família":   pa["Familia"],
            "Categoria": categoria,
            "Código PA": pa_code,
            "Descrição": pa["ItemName"][:45],
            "ABC":       pa["ABC"]
        }

        # Estoque inicial = estoque atual (carteira é abatida POR SEMANA usando ShipDate de cada pedido).
        estoque_pa_sim = pa.get("Estoque", 0) or 0
        consumo_semana = media_dia * 7

        chegadas_pa = chegadas.get(pa_code, {})
        vendas_pa   = vendas_por_pa.get(pa_code, {})

        # inflow_from[i] = total de chegadas disponíveis a partir da semana i (inclusiva)
        # Usado pra decidir ATENÇÃO (atrasado mas será coberto) vs RUPTURA (perdido).
        inflow_from = [0] * (len(semanas) + 1)
        for i in range(len(semanas) - 1, -1, -1):
            inflow_from[i] = inflow_from[i + 1] + chegadas_pa.get(i, 0)

        # Acumula débito não-coberto para sustentar status ATENÇÃO até estoque chegar.
        deficit_acumulado = 0

        for i, semana in enumerate(semanas):
            label = f"S{semana.strftime('%V')}\n{semana.strftime('%d/%m')}"

            chegada_qtd = chegadas_pa.get(i, 0)
            vendas_qtd  = vendas_pa.get(i, 0)
            outflow     = max(consumo_semana, vendas_qtd)  # baseline de média OU pedidos confirmados, o que for maior

            # Chegada paga primeiro o déficit acumulado, sobra vai pro estoque
            chegada_para_estoque = chegada_qtd
            if deficit_acumulado > 0 and chegada_qtd > 0:
                paga = min(chegada_qtd, deficit_acumulado)
                deficit_acumulado     -= paga
                chegada_para_estoque  -= paga

            estoque_pa_sim += chegada_para_estoque

            if estoque_pa_sim >= outflow:
                estoque_pa_sim -= outflow
                shortfall = 0
            else:
                shortfall = outflow - estoque_pa_sim
                deficit_acumulado += shortfall
                estoque_pa_sim = 0

            inflow_futuro = inflow_from[i + 1]  # total que ainda vai chegar depois desta semana

            # Decisão de status (ordem importa)
            if shortfall > 0:
                # Faltou estoque essa semana
                if inflow_futuro >= deficit_acumulado:
                    linha[label] = "ATENÇÃO"   # atrasado mas será coberto
                else:
                    linha[label] = "RUPTURA"   # nada cobre, venda perdida
            elif deficit_acumulado > 0:
                # Não faltou hoje mas ainda há débito de semanas anteriores não pago
                linha[label] = "ATENÇÃO"
            elif chegada_qtd > 0:
                linha[label] = "CHEGADA"
            elif estoque_pa_sim < consumo_semana * 2:
                linha[label] = "ATENÇÃO"        # menos de 2 semanas de cobertura
            else:
                linha[label] = "OK"

        calendario.append(linha)

    df_cal = pd.DataFrame(calendario)
    df_cal = df_cal.sort_values(["Família", "ABC"])
    semana_labels = [f"S{s.strftime('%V')}\n{s.strftime('%d/%m')}" for s in semanas]
    print(f"✅ Calendário gerado para {len(df_cal)} PAs")
    return df_cal, semana_labels

# ============================================================
# 📊  GERAR EXCEL
# ============================================================

def criar_aba_resumo(ws3, df_capacidade, pa_insumos, estoque_dict, pedidos_dict, carteira_dict, nome_dict):
    """
    Cria a 3ª aba com:
      TABELA 1 — PAs CRÍTICOS PARA PRODUZIR (Peças Possíveis = 0)
      TABELA 2 — INSUMOS CRÍTICOS PARA COMPRAR
    """
    # Estilos
    HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
    SUBHEAD_FILL = PatternFill("solid", fgColor="5B9BD5")
    VERMELHO    = PatternFill("solid", fgColor="FFC7CE")
    AMARELO     = PatternFill("solid", fgColor="FFEB9C")
    VERDE       = PatternFill("solid", fgColor="C6EFCE")
    ABC_A       = PatternFill("solid", fgColor="FF0000")
    ABC_B       = PatternFill("solid", fgColor="FFC000")
    ABC_C       = PatternFill("solid", fgColor="92D050")
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    ws3.freeze_panes = "A2"

    # ── TABELA 1 — PAs CRÍTICOS PARA PRODUZIR ─────────────────
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "🚨 PAs CRÍTICOS — Peças Possíveis = 0"
    ws3["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill      = HEADER_FILL
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    headers_t1 = ["Família", "Código PA", "Descrição", "ABC", "Carteira", "Estoque PA", "Insumo Gargalo", "Motivo"]
    for col, h in enumerate(headers_t1, 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = SUBHEAD_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = borda
    ws3.row_dimensions[2].height = 26

    # Filtra PAs críticos
    criticos = df_capacidade[df_capacidade["Peças Possíveis"] == 0].copy()
    criticos["_urg"] = (criticos["Carteira"] > 0).astype(int)
    abc_ord = {"A": 0, "B": 1, "C": 2}
    criticos["_abc_ord"] = criticos["Classe ABC"].map(abc_ord).fillna(3)
    criticos = criticos.sort_values(["_urg", "_abc_ord", "Família"], ascending=[False, True, True])

    r = 3
    for _, row in criticos.iterrows():
        valores = [
            row["Família"], row["Código PA"], row["Descrição PA"], row["Classe ABC"],
            int(row["Carteira"]), int(row["Estoque PA"]),
            row["Insumo Gargalo"], row.get("Insumo Zerado", "") or row.get("Motivo Zero", "")
        ]
        for col_idx, val in enumerate(valores, 1):
            c = ws3.cell(row=r, column=col_idx, value=val)
            c.border    = borda
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if col_idx == 4:  # ABC
                if val == "A":   c.fill = ABC_A; c.font = Font(bold=True, color="FFFFFF")
                elif val == "B": c.fill = ABC_B; c.font = Font(bold=True)
                elif val == "C": c.fill = ABC_C; c.font = Font(bold=True)
            if col_idx == 5 and isinstance(val, (int, float)) and val > row["Estoque PA"]:
                c.fill = VERMELHO
                c.font = Font(bold=True, color="9C0006")
            if col_idx == 8 and val:
                c.fill = VERMELHO
                c.font = Font(color="9C0006", size=9)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws3.row_dimensions[r].height = 18
        r += 1

    # ── TABELA 2 — INSUMOS CRÍTICOS PARA COMPRAR ──────────────
    r += 2  # espaço
    cabecalho_row = r
    ws3.merge_cells(f"A{r}:H{r}")
    c = ws3.cell(row=r, column=1, value="🛒 INSUMOS CRÍTICOS PARA COMPRAR")
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[r].height = 30
    r += 1

    headers_t2 = ["Insumo", "Descrição", "Estoque Atual", "Pedido em Aberto", "Qtd Chegando",
                  "Necessidade (cart + 2m)", "Quanto Comprar", "Data Necessidade"]
    for col, h in enumerate(headers_t2, 1):
        c = ws3.cell(row=r, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = SUBHEAD_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = borda
    ws3.row_dimensions[r].height = 30
    r += 1

    # Índice pa_code → (media_dia, carteira, abc) a partir do df_capacidade
    info_pa = {}
    for _, row in df_capacidade.iterrows():
        info_pa[row["Código PA"]] = {
            "media":    float(row.get("Média/dia", 0) or 0),
            "carteira": int(row.get("Carteira", 0) or 0),
        }

    # Insumos gargalo dos PAs críticos
    insumos_criticos = set(criticos["Insumo Gargalo"].dropna().unique())
    insumos_criticos.discard("")
    insumos_criticos.discard("SEM BOM")

    hoje = datetime.date.today()
    resumo_insumos = []

    for ins in insumos_criticos:
        # Soma necessidade considerando TODOS os PAs que usam esse insumo
        carteira_total   = 0.0
        consumo_2m_total = 0.0
        consumo_diario_total = 0.0
        for pa_code, ins_dict in pa_insumos.items():
            if ins not in ins_dict:
                continue
            qtd_por_pa = ins_dict[ins]
            if qtd_por_pa <= 0:
                continue
            info = info_pa.get(pa_code, {})
            carteira_total   += info.get("carteira", 0) * qtd_por_pa
            consumo_2m_total += info.get("media", 0) * 60 * qtd_por_pa
            consumo_diario_total += info.get("media", 0) * qtd_por_pa

        estoque_atual = estoque_dict.get(ins, 0)
        ped = pedidos_dict.get(ins, {})
        qtd_chegando = ped.get("QtdChegando", 0) or 0
        prox_entrega = ped.get("ProximaEntrega", None)
        pedido_aberto = 1 if qtd_chegando > 0 else 0

        necessidade  = carteira_total + consumo_2m_total
        quanto_comprar = max(0, necessidade - estoque_atual - qtd_chegando)

        if consumo_diario_total > 0:
            dias_ate_rupture = estoque_atual / consumo_diario_total
            data_necessidade = hoje + datetime.timedelta(days=int(dias_ate_rupture))
        else:
            data_necessidade = None

        resumo_insumos.append({
            "ins": ins,
            "nome": str(nome_dict.get(ins, "")).strip(),
            "estoque": estoque_atual,
            "pedido_aberto": pedido_aberto,
            "qtd_chegando": qtd_chegando,
            "necessidade": necessidade,
            "quanto_comprar": quanto_comprar,
            "data_necessidade": data_necessidade,
            "prox_entrega": prox_entrega,
        })

    # Ordena por data de necessidade (None/mais antigo primeiro)
    def _sort_key(x):
        d = x["data_necessidade"]
        return d if d else datetime.date(9999, 12, 31)
    resumo_insumos.sort(key=_sort_key)

    for item in resumo_insumos:
        valores = [
            item["ins"],
            item["nome"][:45],
            int(item["estoque"]),
            "SIM" if item["pedido_aberto"] else "—",
            int(item["qtd_chegando"]),
            int(item["necessidade"]),
            int(item["quanto_comprar"]),
            item["data_necessidade"].strftime("%d/%m/%Y") if item["data_necessidade"] else "—",
        ]
        for col_idx, val in enumerate(valores, 1):
            c = ws3.cell(row=r, column=col_idx, value=val)
            c.border    = borda
            c.alignment = Alignment(horizontal="center", vertical="center")

            if col_idx == 7 and isinstance(val, (int, float)) and val > 0:
                c.fill = AMARELO
                c.font = Font(bold=True)
            if col_idx == 8 and item["data_necessidade"]:
                dias = (item["data_necessidade"] - hoje).days
                if dias <= 7:
                    c.fill = VERMELHO
                    c.font = Font(bold=True, color="9C0006")
                elif dias <= 30:
                    c.fill = AMARELO
                else:
                    c.fill = VERDE
        ws3.row_dimensions[r].height = 18
        r += 1

    # Larguras
    larguras = [14, 42, 14, 14, 14, 18, 16, 16]
    for i, w in enumerate(larguras, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w


def criar_aba_mrp(ws4, df_mrp):
    """Aba 4 — MRP: Sugestão de Ordens de Compra"""
    HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
    SUBHEAD_FILL = PatternFill("solid", fgColor="375623")
    VERMELHO     = PatternFill("solid", fgColor="FFC7CE")
    AMARELO      = PatternFill("solid", fgColor="FFEB9C")
    VERDE        = PatternFill("solid", fgColor="C6EFCE")
    CINZA        = PatternFill("solid", fgColor="F2F2F2")
    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))

    ws4.freeze_panes = "A3"
    n_cols   = len(df_mrp.columns)
    col_last = get_column_letter(n_cols)

    # Título
    ws4.merge_cells(f"A1:{col_last}1")
    ws4["A1"] = f"⚙️  MRP — SUGESTÃO DE COMPRAS — Gerado em {datetime.date.today().strftime('%d/%m/%Y')}"
    ws4["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws4["A1"].fill      = HEADER_FILL
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    # Legenda
    ws4.merge_cells(f"A2:{col_last}2")
    ws4["A2"] = "🔴 URGENTE (já em falta)   🟡 ATENÇÃO (falta em até 7 dias)   🟢 OK (falta em mais de 7 dias)   |   Lead time e data de emissão: a definir após cadastro"
    ws4["A2"].font      = Font(size=10, italic=True, name="Calibri")
    ws4["A2"].fill      = PatternFill("solid", fgColor="EBF0FA")
    ws4["A2"].alignment = Alignment(horizontal="center")
    ws4.row_dimensions[2].height = 20

    # Cabeçalhos
    for col, h in enumerate(df_mrp.columns, 1):
        c = ws4.cell(row=3, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill      = SUBHEAD_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = borda
    ws4.row_dimensions[3].height = 30

    # Dados
    for r, (_, row) in enumerate(df_mrp.iterrows(), 4):
        for ci, (col, val) in enumerate(row.items(), 1):
            c = ws4.cell(row=r, column=ci, value=val)
            c.border    = borda
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if col == "Urgência":
                if "URGENTE" in str(val):
                    c.fill = VERMELHO; c.font = Font(bold=True, size=10, name="Calibri")
                elif "ATENÇÃO" in str(val):
                    c.fill = AMARELO;  c.font = Font(bold=True, size=10, name="Calibri")
                else:
                    c.fill = VERDE
            elif col == "Nec. Líquida" and isinstance(val, (int, float)) and val > 0:
                c.fill = AMARELO; c.font = Font(bold=True, size=10, name="Calibri")
            elif col == "Pedido em Aberto" and val == "SIM":
                c.fill = VERDE
            elif col in ("Lead Time (dias)", "Emitir Pedido Até") and "definir" in str(val):
                c.fill = CINZA
                c.font = Font(italic=True, color="888888", size=9, name="Calibri")
            elif col == "PAs Afetados":
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws4.row_dimensions[r].height = 20

    # Totalizador
    tot_row = len(df_mrp) + 4
    ws4.merge_cells(f"A{tot_row}:H{tot_row}")
    c = ws4.cell(row=tot_row, column=1,
                 value=f"Total de insumos que precisam de reposição: {len(df_mrp)}")
    c.font      = Font(bold=True, size=11, name="Calibri")
    c.fill      = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")

    # Larguras
    larguras = [12, 18, 42, 13, 13, 13, 14, 13, 13, 14, 14, 28, 35]
    for i, w in enumerate(larguras, 1):
        if i <= n_cols:
            ws4.column_dimensions[get_column_letter(i)].width = w


def criar_aba_bom_explodida(ws5, df_pas_abc, df_bom, estoque_dict, nome_dict, grupo_dict, carteira_dict):
    """
    Aba BOM Explodida — estrutura em árvore por PA
    PA (pai) → Conjuntos/PIs (filhos) → Insumos (folhas)
    """
    HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
    PA_FILL      = PatternFill("solid", fgColor="2E75B6")
    PI_FILL      = PatternFill("solid", fgColor="DEEAF1")
    INS_FILL     = PatternFill("solid", fgColor="FFFFFF")
    INS_ALT_FILL = PatternFill("solid", fgColor="F7FBFF")
    VERMELHO     = PatternFill("solid", fgColor="FFC7CE")
    AMARELO      = PatternFill("solid", fgColor="FFEB9C")
    VERDE        = PatternFill("solid", fgColor="C6EFCE")
    ABC_A        = PatternFill("solid", fgColor="FF0000")
    ABC_B        = PatternFill("solid", fgColor="FFC000")
    ABC_C        = PatternFill("solid", fgColor="92D050")

    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))
    borda_grossa = Border(left=Side(style="medium"), right=Side(style="medium"),
                          top=Side(style="medium"),  bottom=Side(style="medium"))

    ws5.freeze_panes = "A3"

    # Monta bom_dict
    bom_dict = {}
    for _, row in df_bom.iterrows():
        f = row["Father"]
        if f not in bom_dict:
            bom_dict[f] = []
        bom_dict[f].append((row["Insumo"], float(row["Quantity"])))

    # Título
    ws5.merge_cells("A1:H1")
    ws5["A1"] = f"BOM EXPLODIDA — Estrutura de Produtos — Gerado em {datetime.date.today().strftime('%d/%m/%Y')}"
    ws5["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws5["A1"].fill      = HEADER_FILL
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 32

    # Legenda
    ws5.merge_cells("A2:H2")
    ws5["A2"] = "Nível 0 = PA   |   Nível 1 = Conjunto/PI direto   |   Nível 2+ = Insumo final   |   Qtd Necessária = por unidade do PA"
    ws5["A2"].font      = Font(size=10, italic=True, name="Calibri")
    ws5["A2"].fill      = PatternFill("solid", fgColor="EBF0FA")
    ws5["A2"].alignment = Alignment(horizontal="center")
    ws5.row_dimensions[2].height = 18

    # Cabeçalhos
    headers = ["Nível", "Código", "Descrição", "ABC / Tipo", "Qtd p/ 1 PA", "Estoque Atual", "Cob. (dias)", "Status"]
    for col, h in enumerate(headers, 1):
        c = ws5.cell(row=3, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = borda
    ws5.row_dimensions[3].height = 28

    r = 4

    # Função recursiva para escrever a árvore
    def escrever_arvore(item_code, qtd_pai, nivel, media_dia_pa):
        nonlocal r

        estoque = estoque_dict.get(item_code, 0)
        nome    = str(nome_dict.get(item_code, "")).strip()
        grupo   = grupo_dict.get(item_code, 0)

        # Tipo do item
        tipo_map = {125: "PA", 123: "Revenda", 126: "Conjunto", 124: "Prod. Interm.",
                    120: "Componente", 121: "Embalagem", 122: "Consumível", 119: "Mat. Prima"}
        tipo = tipo_map.get(grupo, f"Grp {grupo}")

        # Indentação visual
        prefixo = "    " * nivel
        if nivel == 0:
            prefixo = ""
        elif nivel == 1:
            prefixo = "  ├── "
        else:
            prefixo = "  │   " * (nivel - 1) + "  └── "

        # Cobertura em dias
        if media_dia_pa > 0 and qtd_pai > 0:
            consumo_dia_ins = media_dia_pa * qtd_pai
            cob_dias = round(estoque / consumo_dia_ins, 1) if consumo_dia_ins > 0 else "-"
        else:
            cob_dias = "-"

        # Status do estoque
        if estoque == 0 and grupo not in (125, 123, 126, 124):
            status = "🔴 ZERADO"
            est_fill = VERMELHO
        elif isinstance(cob_dias, float) and cob_dias < 7:
            status = "🟡 BAIXO"
            est_fill = AMARELO
        elif isinstance(cob_dias, float) and cob_dias >= 7:
            status = "🟢 OK"
            est_fill = VERDE
        else:
            status = "—"
            est_fill = None

        # Escolhe estilo da linha
        if nivel == 0:
            fill = PA_FILL
            font_color = "FFFFFF"
            bold = True
            font_size = 11
        elif nivel == 1 and grupo in (124, 126):
            fill = PI_FILL
            font_color = "1A3A5C"
            bold = True
            font_size = 10
        else:
            fill = INS_ALT_FILL if r % 2 == 0 else INS_FILL
            font_color = "000000"
            bold = False
            font_size = 10

        # Escreve linha
        valores = [
            f"{'  ' * nivel}{'PA' if nivel==0 else ('PI/CJ' if grupo in (124,126) else 'INS')}",
            item_code,
            prefixo + nome[:50],
            tipo if nivel > 0 else "",
            round(qtd_pai, 4) if nivel > 0 else 1,
            int(estoque),
            cob_dias,
            status if nivel > 0 else "",
        ]

        for ci, val in enumerate(valores, 1):
            c = ws5.cell(row=r, column=ci, value=val)
            c.border    = borda if nivel > 0 else borda_grossa
            c.font      = Font(bold=bold, color=font_color, size=font_size, name="Calibri")
            c.fill      = fill
            c.alignment = Alignment(horizontal="center" if ci not in (3,) else "left",
                                    vertical="center", wrap_text=True)

            # Coluna estoque com cor de status
            if ci == 6 and est_fill and nivel > 0:
                c.fill = est_fill

        ws5.row_dimensions[r].height = 18 if nivel > 0 else 22
        r += 1

        # Filhos
        filhos = bom_dict.get(item_code, [])
        for (filho, qtd_filho) in filhos:
            qtd_acumulada = qtd_pai * qtd_filho if nivel > 0 else qtd_filho
            escrever_arvore(filho, qtd_acumulada, nivel + 1, media_dia_pa)

    # Ordena PAs por família e ABC
    df_sorted = df_pas_abc.sort_values(["Familia", "ABC", "ItemCode"])

    familia_atual = None
    for _, pa in df_sorted.iterrows():
        pa_code   = pa["ItemCode"]
        familia   = pa["Familia"]
        media_dia = pa["MediaDiaria"]
        abc       = pa["ABC"]

        # Só processa PAs que têm BOM
        if pa_code not in bom_dict:
            continue

        # Separador de família
        if familia != familia_atual:
            familia_atual = familia
            ws5.merge_cells(f"A{r}:H{r}")
            c = ws5.cell(row=r, column=1, value=f"  Família: {familia}")
            c.font      = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            c.fill      = PatternFill("solid", fgColor="2E75B6")
            c.alignment = Alignment(vertical="center")
            c.border    = borda_grossa
            ws5.row_dimensions[r].height = 22
            r += 1

        # Linha do PA
        escrever_arvore(pa_code, 1, 0, media_dia)

        # Linha em branco entre PAs
        ws5.row_dimensions[r].height = 6
        r += 1

    # Larguras
    larguras = [8, 18, 55, 14, 12, 14, 12, 14]
    for i, w in enumerate(larguras, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w


def gerar_excel(df_capacidade, df_calendario, semanas,
                pa_insumos=None, estoque_dict=None, pedidos_dict=None,
                carteira_dict=None, nome_dict=None, df_mrp=None,
                df_pas_abc=None, df_bom=None, grupo_dict=None):
    print("📊 Gerando Excel...")
    wb = Workbook()

    VERDE       = PatternFill("solid", fgColor="C6EFCE")
    AMARELO     = PatternFill("solid", fgColor="FFEB9C")
    VERMELHO    = PatternFill("solid", fgColor="FFC7CE")
    AZUL        = PatternFill("solid", fgColor="BDD7EE")
    CINZA_CLARO = PatternFill("solid", fgColor="D9D9D9")
    HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
    FAM_FILL    = PatternFill("solid", fgColor="D9E1F2")
    ABC_A       = PatternFill("solid", fgColor="FF0000")
    ABC_B       = PatternFill("solid", fgColor="FFC000")
    ABC_C       = PatternFill("solid", fgColor="92D050")

    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # ── ABA 1 — CAPACIDADE ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "🏭 Capacidade de Produção"
    ws1.freeze_panes = "B4"

    n_cols = len(df_capacidade.columns)
    col_last = get_column_letter(n_cols)

    # Título
    ws1.merge_cells(f"A1:{col_last}1")
    ws1["A1"] = f"CAPACIDADE DE PRODUÇÃO — Gerado em {datetime.date.today().strftime('%d/%m/%Y')}"
    ws1["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill      = HEADER_FILL
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    # Legenda
    ws1.merge_cells(f"A2:{col_last}2")
    ws1["A2"] = "🟢 Cob. > 15 dias   🟡 Cob. 7-15 dias   🔴 Cob. < 7 dias   |   ABC: 🔴 A (top 80%)   🟡 B (80-95%)   🟢 C (95-100%)   |   Peças calculadas em cascata por prioridade"
    ws1["A2"].font      = Font(size=10, italic=True)
    ws1["A2"].fill      = PatternFill("solid", fgColor="EBF0FA")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 20

    # Cabeçalhos
    headers = list(df_capacidade.columns)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = borda
    ws1.row_dimensions[3].height = 32

    # Dados — agrupa por família
    familia_atual = None
    excel_row = 4
    for _, row in df_capacidade.iterrows():
        familia = row["Família"]

        if familia != familia_atual:
            familia_atual = familia
            ws1.merge_cells(f"A{excel_row}:{col_last}{excel_row}")
            cell = ws1.cell(row=excel_row, column=1, value=f"  📦 Família: {familia}")
            cell.font      = Font(bold=True, size=11, color="1A3A5C")
            cell.fill      = FAM_FILL
            cell.alignment = Alignment(vertical="center")
            ws1.row_dimensions[excel_row].height = 20
            excel_row += 1

        for col_idx, (col, val) in enumerate(row.items(), 1):
            cell = ws1.cell(row=excel_row, column=col_idx, value=val)
            cell.border    = borda
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == "Classe ABC":
                if val == "A":   cell.fill = ABC_A; cell.font = Font(bold=True, color="FFFFFF")
                elif val == "B": cell.fill = ABC_B; cell.font = Font(bold=True)
                elif val == "C": cell.fill = ABC_C; cell.font = Font(bold=True)

            elif col in ("Cob. PA (dias)", "Cob. Insumo (dias)") and val != "-":
                try:
                    v = float(val)
                    if v < 7:    cell.fill = VERMELHO
                    elif v < 15: cell.fill = AMARELO
                    else:        cell.fill = VERDE
                except: pass

            elif col == "Carteira" and isinstance(val, (int, float)):
                estoque_pa = row.get("Estoque PA", 0)
                if val > estoque_pa:
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    cell.font = Font(bold=True, color="C00000")

            elif col == "Peças Possíveis" and isinstance(val, (int, float)):
                if val == 0:
                    cell.fill = VERMELHO
                    cell.font = Font(bold=True, color="9C0006")
                elif val > 0:
                    cell.fill = VERDE
                    cell.font = Font(bold=True, color="375623")

            elif col == "Insumo Zerado" and val:
                cell.fill      = PatternFill("solid", fgColor="FCE4D6")
                cell.font      = Font(bold=True, color="9C0006", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws1.row_dimensions[excel_row].height = 18
        excel_row += 1

    # Larguras aba 1
    larguras = [14, 22, 45, 8, 10, 10, 9, 10, 18, 12, 12, 13, 12, 12, 50]
    for i, w in enumerate(larguras, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── ABA 2 — CALENDÁRIO ──────────────────────────────────
    ws2 = wb.create_sheet("📅 Calendário de Ruptura")
    ws2.freeze_panes = "E3"

    total_cols = 4 + len(semanas)

    ws2.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws2["A1"] = f"CALENDÁRIO DE RUPTURA SEMANAL — {datetime.date.today().year}"
    ws2["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill      = HEADER_FILL
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    for col, h in enumerate(["Família", "Código PA", "Descrição", "ABC"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = borda

    for i, s in enumerate(semanas, 5):
        cell = ws2.cell(row=2, column=i, value=s.replace("\n", " "))
        cell.font      = Font(bold=True, color="FFFFFF", size=8)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = borda
    ws2.row_dimensions[2].height = 32

    familia_atual = None
    cal_row = 3
    for _, row in df_calendario.iterrows():
        familia = row["Família"]

        if familia != familia_atual:
            familia_atual = familia
            ws2.merge_cells(f"A{cal_row}:{get_column_letter(total_cols)}{cal_row}")
            cell = ws2.cell(row=cal_row, column=1, value=f"  📦 Família: {familia}")
            cell.font      = Font(bold=True, size=11, color="1A3A5C")
            cell.fill      = FAM_FILL
            cell.alignment = Alignment(vertical="center")
            ws2.row_dimensions[cal_row].height = 20
            cal_row += 1

        ws2.cell(row=cal_row, column=1, value=row["Família"]).border  = borda
        ws2.cell(row=cal_row, column=2, value=row["Código PA"]).border = borda
        ws2.cell(row=cal_row, column=3, value=row["Descrição"]).border = borda

        abc_cell = ws2.cell(row=cal_row, column=4, value=row["ABC"])
        abc_cell.border    = borda
        abc_cell.alignment = Alignment(horizontal="center")
        if row["ABC"] == "A":   abc_cell.fill = ABC_A; abc_cell.font = Font(bold=True, color="FFFFFF")
        elif row["ABC"] == "B": abc_cell.fill = ABC_B; abc_cell.font = Font(bold=True)
        elif row["ABC"] == "C": abc_cell.fill = ABC_C; abc_cell.font = Font(bold=True)

        for i, s in enumerate(semanas, 5):
            val  = row.get(s, "OK")
            cell = ws2.cell(row=cal_row, column=i, value="")
            cell.border    = borda
            cell.alignment = Alignment(horizontal="center")
            if val == "OK":             cell.fill = VERDE
            elif val == "ATENÇÃO":      cell.fill = AMARELO
            elif val == "RUPTURA":      cell.fill = VERMELHO
            elif val == "CHEGADA":      cell.fill = AZUL
            elif val == "SEM HISTORICO": cell.fill = CINZA_CLARO

        ws2.row_dimensions[cal_row].height = 16
        cal_row += 1

    leg_row = cal_row + 1
    ws2.merge_cells(f"A{leg_row}:{get_column_letter(total_cols)}{leg_row}")
    leg = ws2.cell(row=leg_row, column=1,
        value="🟢 OK (estoque > demanda semanal)   🟡 ATENÇÃO (estoque < demanda)   🔴 RUPTURA (estoque zerado)   🔵 CHEGADA (pedido de compra chegando)   ⬜ SEM HISTÓRICO (sem consumo nos últimos 3 meses)")
    leg.font      = Font(bold=True, size=10)
    leg.alignment = Alignment(horizontal="center")
    leg.fill      = PatternFill("solid", fgColor="EBF0FA")

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 45
    ws2.column_dimensions["D"].width = 6
    for i in range(5, 5 + len(semanas)):
        ws2.column_dimensions[get_column_letter(i)].width = 8

    # ── ABA 3 — RESUMO ──────────────────────────────────────
    if pa_insumos is not None and estoque_dict is not None:
        ws3 = wb.create_sheet("🚨 Resumo Crítico")
        criar_aba_resumo(ws3, df_capacidade, pa_insumos, estoque_dict,
                         pedidos_dict or {}, carteira_dict or {}, nome_dict or {})

    # ── ABA 4 — MRP ─────────────────────────────────────────
    if df_mrp is not None and len(df_mrp) > 0:
        ws4 = wb.create_sheet("⚙️ MRP — Compras")
        criar_aba_mrp(ws4, df_mrp)

    # ── ABA 5 — BOM EXPLODIDA ────────────────────────────────
    if df_pas_abc is not None and df_bom is not None and grupo_dict is not None:
        ws5 = wb.create_sheet("🏗️ BOM Explodida")
        criar_aba_bom_explodida(ws5, df_pas_abc, df_bom, estoque_dict or {},
                                nome_dict or {}, grupo_dict, carteira_dict or {})

    wb.save(OUTPUT_PATH)
    print(f"✅ Excel salvo em: {OUTPUT_PATH}")

# ============================================================
# 🚀  EXECUÇÃO PRINCIPAL
# ============================================================

def main():
    print(f"\n{'='*55}")
    print(f"🚀 Capacidade de Produção — {datetime.datetime.now():%d/%m/%Y %H:%M}")
    print(f"{'='*55}\n")

    conn = conectar()

    df_pas         = buscar_pas_ativos(conn)
    df_insumos     = buscar_estoque_insumos(conn)
    df_bom         = buscar_bom(conn)
    df_consumo     = buscar_consumo_3m(conn)
    df_carteira    = buscar_carteira_por_pa(conn)
    df_pedidos     = buscar_pedidos_compra(conn)
    df_ped_futuros = buscar_todos_pedidos_futuros(conn)
    df_ordens      = buscar_ordens_producao(conn)

    conn.close()

    df_pas_abc      = calcular_curva_abc(df_pas, df_consumo)
    (df_capacidade, pa_insumos, estoque_dict, pedidos_dict,
     carteira_dict, nome_dict) = calcular_capacidade(df_pas_abc, df_bom, df_insumos, df_pedidos, df_carteira)
    grupo_dict = df_insumos.set_index('ItemCode')['ItmsGrpCod'].to_dict()
    df_cal, semanas = calcular_calendario(df_pas_abc, df_bom, df_insumos, df_ped_futuros)

    # Fase 2 — MRP
    df_mrp = calcular_mrp(df_pas_abc, pa_insumos, estoque_dict, pedidos_dict,
                          carteira_dict, nome_dict, df_ordens)

    gerar_excel(df_capacidade, df_cal, semanas,
                pa_insumos=pa_insumos, estoque_dict=estoque_dict,
                pedidos_dict=pedidos_dict, carteira_dict=carteira_dict,
                nome_dict=nome_dict, df_mrp=df_mrp,
                df_pas_abc=df_pas_abc, df_bom=df_bom, grupo_dict=grupo_dict)

    print(f"\n✅ Concluído! Arquivo: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()